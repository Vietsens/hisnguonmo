# -*- coding: utf-8 -*-
"""Sinh template Excel mẫu cho biểu in Mps000518 - Biên bản/Hợp đồng cung ứng thuốc, vật tư.

Cú pháp token MPS (Inventec.Common.FlexCelExport):
  - Key đơn:  <#KEY;>
  - Band/list: <#Mety.FIELD;> , <#Maty.FIELD;> , <#Mety.#rowpos> (STT)

Lưu ý: các cột band (<#Mety.* / <#Maty.*) bám theo cột view V_HIS_MEDI_CONTACT_METY /
V_HIS_MEDI_CONTACT_MATY (chưa có trong EFMODEL hiện tại). Khi backend bổ sung view thật,
chỉ cần đổi tên field trong file này cho khớp tên cột thực tế rồi sinh lại template.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="D9E1F2")
hdr_font = Font(bold=True, size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
title_font = Font(bold=True, size=14)
italic = Font(italic=True, size=10)
bold = Font(bold=True, size=10)

wb = Workbook()
ws = wb.active
ws.title = "Sheet"

# ----- Thông tin bệnh viện (key đơn do base SetCommonSingleKey sinh tự động) -----
ws.cell(row=1, column=1, value="<#PARENT_ORGANIZATION_NAME;>").font = bold
ws.cell(row=2, column=1, value="<#ORGANIZATION_NAME;>").font = bold
ws.cell(row=3, column=1, value="<#ORGANIZATION_ADDRESS;>").font = italic

# ----- Tiêu đề -----
t = ws.cell(row=5, column=1, value="BIÊN BẢN CUNG ỨNG THUỐC, VẬT TƯ")
t.font = title_font
t.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)

# ----- Thông tin chung hợp đồng (V_HIS_MEDICAL_CONTACT - key đơn reflection) -----
ws.cell(row=7, column=1, value="Số hợp đồng: <#MEDICAL_CONTACT_CODE;>").font = italic
ws.cell(row=8, column=1, value="Tên hợp đồng: <#MEDICAL_CONTACT_NAME;>").font = italic

# ----- Thông tin nhà cung cấp (HIS_SUPPLIER - key đơn reflection) -----
ws.cell(row=10, column=1, value="Nhà cung cấp: <#SUPPLIER_NAME;>   -   Mã: <#SUPPLIER_CODE;>").font = italic
ws.cell(row=11, column=1, value="Địa chỉ: <#ADDRESS;>   -   MST: <#TAX_CODE;>").font = italic
ws.cell(row=12, column=1, value="Người đại diện: <#REPRESENTATIVE;>   -   Chức vụ: <#POSITION;>").font = italic
ws.cell(row=13, column=1, value="Giấy ủy quyền số: <#AUTH_LETTER_NUM;>, <#AUTH_LETTER_ISSUE_DATE_STR;>").font = italic

# ====================== BẢNG THUỐC (band Mety) ======================
ws.cell(row=15, column=1, value="I. DANH SÁCH THUỐC").font = bold

METY_COLS = [
    ("STT",        "<#Mety.#rowpos>"),
    ("Mã thuốc",   "<#Mety.MEDICINE_TYPE_CODE;>"),
    ("Tên thuốc",  "<#Mety.MEDICINE_TYPE_NAME;>"),
    ("Hoạt chất",  "<#Mety.ACTIVE_INGR_BHYT_NAME;>"),
    ("Hàm lượng",  "<#Mety.CONCENTRA;>"),
    ("Dạng BC",    "<#Mety.DOSAGE_FORM;>"),
    ("Hãng SX",    "<#Mety.MANUFACTURER_NAME;>"),
    ("Nước SX",    "<#Mety.NATIONAL_NAME;>"),
    ("Số lượng",   "<#Mety.AMOUNT;>"),
    ("Đơn giá",    "<#Mety.CONTRACT_PRICE;>"),
    ("VAT(%)",     "<#Mety.IMP_VAT_RATIO;>"),
    ("Thành tiền", "<#Mety.VIR_CONTRACT_PRICE;>"),
]
HDR1, DATA1 = 16, 17
for idx, (code, token) in enumerate(METY_COLS, start=1):
    h = ws.cell(row=HDR1, column=idx, value=code)
    h.font = hdr_font; h.fill = hdr_fill; h.alignment = center; h.border = border
    d = ws.cell(row=DATA1, column=idx, value=token)
    d.border = border; d.font = Font(size=10)
    d.alignment = center if idx in (1, 6, 7, 8, 9, 10) else left

# ====================== BẢNG VẬT TƯ (band Maty) ======================
ws.cell(row=19, column=1, value="II. DANH SÁCH VẬT TƯ").font = bold

MATY_COLS = [
    ("STT",        "<#Maty.#rowpos>"),
    ("Mã vật tư",  "<#Maty.MATERIAL_TYPE_CODE;>"),
    ("Tên vật tư", "<#Maty.MATERIAL_TYPE_NAME;>"),
    ("Quy cách",   "<#Maty.CONCENTRA;>"),
    ("Hãng SX",    "<#Maty.MANUFACTURER_NAME;>"),
    ("Nước SX",    "<#Maty.NATIONAL_NAME;>"),
    ("Số lượng",   "<#Maty.AMOUNT;>"),
    ("Đơn giá",    "<#Maty.CONTRACT_PRICE;>"),
    ("VAT(%)",     "<#Maty.IMP_VAT_RATIO;>"),
    ("Thành tiền", "<#Maty.VIR_CONTRACT_PRICE;>"),
]
HDR2, DATA2 = 20, 21
for idx, (code, token) in enumerate(MATY_COLS, start=1):
    h = ws.cell(row=HDR2, column=idx, value=code)
    h.font = hdr_font; h.fill = hdr_fill; h.alignment = center; h.border = border
    d = ws.cell(row=DATA2, column=idx, value=token)
    d.border = border; d.font = Font(size=10)
    d.alignment = center if idx in (1, 6, 7, 8, 9, 10) else left

# ----- Tổng tiền (số + chữ) -----
tot = ws.cell(row=23, column=1, value="Tổng cộng:")
tot.font = bold
ws.cell(row=23, column=12, value="<#SUM_CONTACT_PRICE;>").font = bold
ws.cell(row=23, column=12).alignment = center
ws.cell(row=24, column=1, value="Bằng chữ: <#SUM_CONTACT_PRICE_TEXT;>").font = italic

# ----- Chữ ký + ngày in -----
ws.cell(row=26, column=9, value="<#CURRENT_DATE_SEPARATE_STR;>").font = italic
ws.cell(row=27, column=2, value="ĐẠI DIỆN NHÀ CUNG CẤP").font = bold
ws.cell(row=27, column=9, value="ĐẠI DIỆN BỆNH VIỆN").font = bold

# Độ rộng cột
widths = [6, 12, 26, 20, 11, 10, 16, 12, 9, 12, 8, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[ws.cell(row=16, column=i).column_letter].width = w

out = "Mps000518_BienBanCungUngThuocVatTu__01.xlsx"
wb.save(out)
print("Saved:", out)
