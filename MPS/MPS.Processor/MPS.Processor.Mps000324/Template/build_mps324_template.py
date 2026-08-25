# -*- coding: utf-8 -*-
"""Sinh mau FlexCel Mps000324 - Phieu thanh quyet toan phau thuat / thu thuat."""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

OUT = sys.argv[1]

FONT = "Times New Roman"


def F(size=11, bold=False, italic=False):
    return Font(name=FONT, size=size, bold=bold, italic=italic)


thin = Side(style="thin", color="000000")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

L = Alignment(horizontal="left", vertical="center", wrap_text=True)
LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
R = Alignment(horizontal="right", vertical="center", wrap_text=True)
RT = Alignment(horizontal="right", vertical="top", wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# --- Do rong cot: 7 cot = 7 cot cua bang chi tiet -------------------------
WIDTHS = {
    "A": 5.5,    # STT
    "B": 42.0,   # Ten thuoc va dung cu
    "C": 9.0,    # Don vi
    "D": 10.0,   # So luong
    "E": 13.0,   # Don gia
    "F": 14.0,   # Thanh tien
    "G": 10.0,   # Ghi chu
}
for col, w in WIDTHS.items():
    ws.column_dimensions[col].width = w


def put(ref, value, font=None, align=None, border=None, fmt=None):
    c = ws[ref]
    c.value = value
    c.font = font or F()
    c.alignment = align or L
    if border:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c


def merge(ref):
    ws.merge_cells(ref)
    # openpyxl chi giu style o o goc -> style vien phai set thu cong cho ca dai
    return ref


def box_range(ref):
    """Ke vien cho toan bo o trong vung (ke ca o bi merge)."""
    for row in ws[ref]:
        for c in row:
            c.border = BOX


# =========================================================================
# 1. Tieu de don vi + barcode
# =========================================================================
merge("A1:C1")
put("A1", "<#PARENT_ORGANIZATION_NAME;>", F(11), L)

merge("A2:C2")
put("A2", "<#ORGANIZATION_NAME;>", F(11, bold=True), L)

# Barcode so vao vien - anh duoc ProcessBarCodeTag do vao o nay
merge("F1:G2")
put("F1", "<#BARCODE_IN_CODE_STR;>", F(9), C)

ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 22

# =========================================================================
# 2. Tieu de phieu
# =========================================================================
merge("A4:G4")
put("A4", "PHIẾU THANH TOÁN PHẪU THUẬT - THỦ THUẬT", F(15, bold=True), C)
ws.row_dimensions[4].height = 26

# =========================================================================
# 3. Khoi thong tin hanh chinh
# =========================================================================
merge("A6:C6")
put("A6", "Họ tên người bệnh : <#VIR_PATIENT_NAME;>", F(11), L)
merge("D6:E6")
put("D6", "Tuổi : <#AGE;>", F(11), L)
merge("F6:G6")
put("F6", "Giới tính : <#GENDER_NAME;>", F(11), L)

merge("A7:D7")
put("A7", "Khoa : <#REQUEST_DEPARTMENT_NAME;>", F(11), L)
merge("E7:G7")
put("E7", "Phòng/Giường : <#BED_ROOM_BED_STR;>", F(11), L)

merge("A8:D8")
put("A8", "Thẻ BHYT : <#TDL_HEIN_CARD_NUMBER;>", F(11), L)
merge("E8:G8")
put("E8", "Đối tượng : <#PATIENT_TYPE_NAME;>", F(11), L)

merge("A9:G9")
put("A9", "Chẩn đoán trước mổ : <#if(<#BEFORE_PTTT_ICD_TEXT;>=\"\";<#BEFORE_PTTT_ICD_NAME;>;<#BEFORE_PTTT_ICD_TEXT;>)>", F(11), L)

merge("A10:G10")
put("A10", "Chẩn đoán sau mổ : <#if(<#AFTER_PTTT_ICD_TEXT;>=\"\";<#AFTER_PTTT_ICD_NAME;>;<#AFTER_PTTT_ICD_TEXT;>)>", F(11), L)

merge("A11:G11")
put("A11", "Phẫu thuật / Thủ thuật lúc : <#START_TIME_SEPARATE_STR;>", F(11), L)

merge("A12:C12")
put("A12", "Số phiếu : <#TICKET_NUMBER_STR;>", F(11), L)
merge("D12:G12")
put("D12", "Ghi chú : <#PTTT_NOTE_STR;>", F(11), L)

merge("A13:G13")
put("A13", "<#MAIN_SERVICE_NAME_STR;>", F(11, bold=True), L)

merge("A14:G14")
put("A14", "Phương pháp thực tế : <#REAL_PTTT_METHOD_STR;>", F(11), L)

merge("A15:C15")
put("A15", "Loại phẫu thuật : <#PTTT_GROUP_NAME;>", F(11), L)
merge("D15:G15")
put("D15", "Phương pháp vô cảm : <#EMOTIONLESS_METHOD_NAME;>", F(11), L)

for r in range(6, 16):
    ws.row_dimensions[r].height = 17

# =========================================================================
# 4. Khoi kip mo - lay TU DANH MUC HIS_EXECUTE_ROLE (dataset EkipRolesUsed)
#    1 dong / 1 vai tro. Nhan vai tro do danh muc quyet dinh.
# =========================================================================
merge("A17:B17")
put("A17", "<#EkipRolesUsed.EXECUTE_ROLE_NAME;>", F(11), L)
merge("C17:G17")
put("C17", ": <#EkipRolesUsed.USERNAMES;><#Row Height(Autofit)>", F(11), L)
ws.row_dimensions[17].height = 17

# =========================================================================
# 5. Bang chi tiet
# =========================================================================
HEAD = 19
headers = [
    ("A", "STT"),
    ("B", "Tên thuốc và dụng cụ"),
    ("C", "Đơn vị"),
    ("D", "Số lượng"),
    ("E", "Đơn giá"),
    ("F", "Thành tiền"),
    ("G", "Ghi chú"),   # hao phi -> "Hao Phi"; con lai -> doi tuong thanh toan
]
fill = PatternFill("solid", fgColor="F2F2F2")
for col, text in headers:
    c = put("%s%d" % (col, HEAD), text, F(11, bold=True), C, BOX)
    c.fill = fill
ws.row_dimensions[HEAD].height = 32

# --- Band nhom (master) ---
GRP = HEAD + 1
put("A%d" % GRP, "<#Groups.NUM_ORDER_ROMAN;>", F(11, bold=True), C, BOX)
merge("B%d:E%d" % (GRP, GRP))
put("B%d" % GRP, "<#Groups.SERVICE_TYPE_NAME;>", F(11, bold=True), L, BOX)
put("F%d" % GRP, "<#Groups.TOTAL_AMOUNT;>", F(11, bold=True), R, BOX, "#,##0")
put("G%d" % GRP, "", F(11), C, BOX)
box_range("A%d:G%d" % (GRP, GRP))
ws.row_dimensions[GRP].height = 18

# --- Band dong chi tiet (detail) ---
DET = GRP + 1
put("A%d" % DET, "<#Items.NUM_ORDER;>", F(11), C, BOX)
put("B%d" % DET, "<#Items.SERVICE_NAME;><#Row Height(Autofit)>", F(11), LT, BOX)
put("C%d" % DET, "<#Items.SERVICE_UNIT_NAME;>", F(11), C, BOX)
put("D%d" % DET, "<#Items.AMOUNT;>", F(11), R, BOX, "#,##0.00")
put("E%d" % DET, "<#Items.PRICE;>", F(11), R, BOX, "#,##0")
put("F%d" % DET, "<#Items.INTO_MONEY;>", F(11), R, BOX, "#,##0")
put("G%d" % DET, "<#Items.NOTE;>", F(11), C, BOX)
ws.row_dimensions[DET].height = 17

# --- Dong tong cong ---
TOT = DET + 1
merge("A%d:E%d" % (TOT, TOT))
put("A%d" % TOT, "Tổng chi phí các khoản (Cộng I + II + III .. ) :", F(11, bold=True), R, BOX)
put("F%d" % TOT, "<#GRAND_TOTAL_AMOUNT;>", F(11, bold=True), R, BOX, "#,##0")
put("G%d" % TOT, "", F(11), C, BOX)
box_range("A%d:G%d" % (TOT, TOT))
ws.row_dimensions[TOT].height = 20

# =========================================================================
# 6. Ngay thang + khoi chu ky
# =========================================================================
DATE = TOT + 2
merge("E%d:G%d" % (DATE, DATE))
put("E%d" % DATE, "<#CURRENT_DATE_SEPARATE_STR;>", F(11, italic=True), C)
ws.row_dimensions[DATE].height = 18

SIGN = DATE + 1
sign_cells = [
    ("A", "Vòng ngoài"),
    ("C", "Kỹ thuật viên"),
    ("E", "BS Gây mê"),
    ("F", "BS Trưởng kíp mổ"),
]
put("A%d" % SIGN, "Vòng ngoài", F(11, bold=True), C)
merge("A%d:B%d" % (SIGN, SIGN))
put("C%d" % SIGN, "Kỹ thuật viên", F(11, bold=True), C)
merge("C%d:D%d" % (SIGN, SIGN))
put("E%d" % SIGN, "BS Gây mê", F(11, bold=True), C)
put("F%d" % SIGN, "BS Trưởng kíp mổ", F(11, bold=True), C)
merge("F%d:G%d" % (SIGN, SIGN))
ws.row_dimensions[SIGN].height = 18

# Ten duoi o ky. Chi PTV chinh la xac dinh duoc tu du lieu (co IS_SURG_MAIN
# cua HIS_EXECUTE_ROLE) nen dien san. 3 o con lai de trong cho ky tay -
# khi chot duoc ma vai tro thi thay bang <#USERNAMES_EXECUTE_ROLE_{ma};>
NAME = SIGN + 3
merge("F%d:G%d" % (NAME, NAME))
put("F%d" % NAME, "<#SURG_MAIN_USERNAME_STR;>", F(11, bold=True), C)
for r in range(SIGN + 1, SIGN + 4):
    ws.row_dimensions[r].height = 18

FOOTNOTE = SIGN + 5
merge("A%d:G%d" % (FOOTNOTE, FOOTNOTE))
put("A%d" % FOOTNOTE,
    "Người đăng nhập : <#CURRENT_LOGINNAME;> - <#CURRENT_USERNAME;>",
    F(9, italic=True), L)

# =========================================================================
# 7. Page setup - A4 dung, lap dong tieu de bang, danh so trang
# =========================================================================
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

ws.page_margins.left = 0.4
ws.page_margins.right = 0.3
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.5
ws.page_margins.header = 0.2
ws.page_margins.footer = 0.2

# Lap dong tieu de bang tren moi trang
ws.print_title_rows = "%d:%d" % (HEAD, HEAD)
ws.print_area = "A1:G%d" % FOOTNOTE

# -------------------------------------------------------------------------
# Defined name __Dataset__ : BAT BUOC de FlexCel long master-detail.
# Vung master PHAI bao trum vung detail (giong mau Mps000002 dang chay that).
# Mau Mps000324 cu thieu __ServiceTypes__ nen nhom khong long duoc.
# -------------------------------------------------------------------------
from openpyxl.workbook.defined_name import DefinedName

RANGES = {
    "__EkipRolesUsed__": "Sheet1!$A$17:$G$17",
    "__Groups__": "Sheet1!$A$%d:$G$%d" % (GRP, DET),   # master bao trum detail
    "__Items__": "Sheet1!$A$%d:$G$%d" % (DET, DET),    # detail long ben trong
}
for name, ref in RANGES.items():
    wb.defined_names[name] = DefinedName(name, attr_text=ref)

# Footer: Toc do an toan - dung ma Excel goc (&P/&N/&T), khong phu thuoc FlexCel
ws.oddFooter.left.text = "Tờ : [&P - &N]"
ws.oddFooter.left.size = 8
ws.oddFooter.left.font = FONT
ws.oddFooter.right.text = "&T"
ws.oddFooter.right.size = 8
ws.oddFooter.right.font = FONT

wb.save(OUT)
print("saved:", OUT)
print("HEAD row =", HEAD, "| GROUP band =", GRP, "| DETAIL band =", DET, "| TOTAL =", TOT)
